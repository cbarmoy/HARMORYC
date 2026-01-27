import asyncio
import json
import os
import platform
import random
import re
import threading
import time
import uuid
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple, Union

import flet as ft

try:
    import websockets
except Exception:  # pragma: no cover
    websockets = None  # type: ignore

try:
    import pandas as pd
    import matplotlib.pyplot as plt
    import matplotlib.dates as mdates
    from openpyxl import load_workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    HAS_EXCEL_SUPPORT = True
except Exception:
    HAS_EXCEL_SUPPORT = False


# -------------------------------
# Utilities
# -------------------------------

APP_NAME = "Questionnaire Mémoire Épisodique"
EXPERIMENT_FILE = "experiment_data.json"
SESSIONS_DIR = "sessions"
ASSETS_DIR = "assets"
WS_PORT = 8765

# Randomization control: "fixed" (common order) or "per_session" (different each session)
RANDOM_MODE = "fixed"
RANDOM_SEED = "HARMORYC_V2"

V2_ASSETS_SUBDIR = "HARMORYC_VR_images_rappels"
IMAGE_EXTS = {".jpg", ".jpeg", ".png", ".gif", ".webp"}
ROOM_GRID_ORDER = ["room8", "room5", "room3", "room7", "room4", "room6", "room10", "room9", "room2", "room1"]

# Durations (ms)
STEP1_DURATION_MS = 10000  # Étape I
STEP2_DURATION_MS = 20000  # Étape II (where/when)
STEP3_DURATION_MS = 10000  # Étape III
STEP4_DURATION_MS = 10000  # Étape IV
RAPPEL_IMMEDIAT_DURATION_MS = 10000  # Rappel immédiat
INTER_DELAY_MS = 3000      # Écran uniforme entre questions

# Image heights by question type (uniform per type)
IMAGE_HEIGHT_STEP1 = 320       # Étape I - what?
IMAGE_HEIGHT_STEP2 = 280       # Étape II - where/when
IMAGE_HEIGHT_STEP3 = 250       # Étape III - what-where-when
IMAGE_HEIGHT_STEP4 = 320       # Étape IV - test ultime
IMAGE_HEIGHT_RAPPEL = 320      # Rappel immédiat


def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def safe_filename(s: str) -> str:
    return "".join(c if c.isalnum() or c in ("-", "_", ".", " ") else "_" for c in s).strip()


def ensure_sample_experiment_exists() -> None:
    """Creates a sample experiment_data.json with placeholders if missing."""
    if Path(EXPERIMENT_FILE).exists():
        return

    Path(ASSETS_DIR, "rooms").mkdir(parents=True, exist_ok=True)
    Path(ASSETS_DIR, "objects").mkdir(parents=True, exist_ok=True)
    Path(ASSETS_DIR, "composites").mkdir(parents=True, exist_ok=True)

    rooms = [
        {"id": f"room{i}", "name": f"Salle {i}", "image": f"{ASSETS_DIR}/rooms/room{i}.jpg"}
        for i in range(1, 11)
    ]

    familiar_objects = []
    for i in range(1, 11):
        familiar_objects.append({
            "id": f"obj{i}",
            "name": f"Objet {i}",
            "image": f"{ASSETS_DIR}/objects/obj{i}.jpg",
            "room_id": f"room{i}",
            "timing": "jour" if i % 2 else "nuit",
        })

    new_objects = [
        {"id": f"new{i}", "name": f"Nouvel objet {i}", "image": f"{ASSETS_DIR}/objects/new{i}.jpg"}
        for i in range(1, 11)
    ]

    step4_trials = []
    for i in range(1, 21):
        step4_trials.append({
            "id": f"step4_{i}",
            "image": f"{ASSETS_DIR}/composites/step4_{i}.jpg",
            "is_correct": True if i % 2 else False,
            "description": f"Scène {i}"
        })

    rappel_immediat_trials = []
    for i in range(1, 7):
        rappel_immediat_trials.append({
            "id": f"rappel_immediat_{i}",
            "image": f"{ASSETS_DIR}/composites/rappel_immediat_{i}.jpg",
            "is_correct": True if i % 2 else False,
            "description": f"Scène RI {i}"
        })

    sample = {
        "experiment_name": "MemoireEpisodique_v1",
        "rooms": rooms,
        "familiar_objects": familiar_objects,
        "new_objects": new_objects,
        "step4_trials": step4_trials,
        "rappel_immediat_trials": rappel_immediat_trials,
    }
    with open(EXPERIMENT_FILE, "w", encoding="utf-8") as f:
        json.dump(sample, f, indent=2, ensure_ascii=False)


# -------------------------------
# VR WebSocket Server
# -------------------------------

class VRWebSocketServer:
    def __init__(self, host: str = "0.0.0.0", port: int = WS_PORT):
        self.host = host
        self.port = port
        self._loop: Optional[asyncio.AbstractEventLoop] = None
        self._thread: Optional[threading.Thread] = None
        self._server: Optional[asyncio.base_events.Server] = None
        self._running = threading.Event()
        self._messages: List[Dict[str, Any]] = []
        self._lock = threading.Lock()

    async def _handler(self, websocket, _path):
        async for message in websocket:
            received_perf = time.perf_counter()
            received_at = now_iso()
            payload: Union[Dict[str, Any], str]
            try:
                payload = json.loads(message)
            except Exception:
                payload = str(message)
            with self._lock:
                self._messages.append({
                    "received_at": received_at,
                    "received_perf": received_perf,
                    "payload": payload,
                })

    def start(self) -> None:
        if self._running.is_set():
            return
        if websockets is None:
            return

        def _run():
            try:
                self._loop = asyncio.new_event_loop()
                asyncio.set_event_loop(self._loop)

                async def start_server():
                    return await websockets.serve(self._handler, self.host, self.port)

                self._server = self._loop.run_until_complete(start_server())
                self._running.set()
                self._loop.run_forever()
            finally:
                try:
                    if self._server is not None:
                        self._server.close()
                        self._loop and self._loop.run_until_complete(self._server.wait_closed())
                finally:
                    self._running.clear()
                    try:
                        self._loop and self._loop.close()
                    except Exception:
                        pass

        self._thread = threading.Thread(target=_run, name="VRWebSocketServer", daemon=True)
        self._thread.start()

    def stop(self) -> None:
        if not self._running.is_set():
            return
        if self._loop is not None:
            self._loop.call_soon_threadsafe(self._loop.stop)
        if self._thread is not None:
            self._thread.join(timeout=2.0)
        self._running.clear()

    def drain_messages(self) -> List[Dict[str, Any]]:
        with self._lock:
            msgs = list(self._messages)
            self._messages.clear()
            return msgs

    def is_running(self) -> bool:
        return self._running.is_set()


# -------------------------------
# Flet App
# -------------------------------

class ExperimentApp:
    def __init__(self, page: ft.Page):
        self.page = page
        self.page.title = APP_NAME
        self.page.window_min_width = 900
        self.page.window_min_height = 650
        self.page.padding = 0
        self.page.spacing = 0
        self.page.theme_mode = ft.ThemeMode.LIGHT
        self.page.assets_dir = ASSETS_DIR
        self.page.bgcolor = ft.Colors.GREY_200

        # Splash screen state
        self.splash_phase: int = 0  # 0=logos, 1=harmoryc, 2=app
        
        # State
        self.data: Dict[str, Any] = {}
        self.experiment_name: str = ""
        self.session_id: str = ""
        self.subject_id: str = ""
        self.task_mode: str = "recall"  # "recall" or "rappel_immediat"
        self.session_started_at: Optional[str] = None
        self.session_ended_at: Optional[str] = None
        self.session_start_perf: Optional[float] = None
        self.running: bool = False

        self.task_queue: List[Dict[str, Any]] = []
        self.current_task_index: int = -1
        self.current_task: Optional[Dict[str, Any]] = None
        self.task_started_at_iso: Optional[str] = None
        self.task_start_perf: Optional[float] = None
        self.task_duration_ms: int = 0
        self.task_active: bool = False
        self.intermission_active: bool = False
        self.intermission_start_perf: Optional[float] = None

        self.selected_choice_id: Optional[str] = None
        self.choice_buttons: Dict[str, ft.Control] = {}
        self.default_image_counter: int = 0  # For alternating default images
        self.nav_drawer_open: bool = False
        self.learning_room_order: List[str] = []
        self.iia_room_order: List[str] = []

        self.records: List[Dict[str, Any]] = []
        self.stage_scores: Dict[str, Dict[str, int]] = {
            "I": {"total": 0, "correct": 0},
            "IIA": {"total": 0, "correct": 0},
            "IIB": {"total": 0, "correct": 0},
            "III": {"total": 0, "correct": 0},
            "IV": {"total": 0, "correct": 0},
            "V": {"total": 0, "correct": 0},
            "RI": {"total": 0, "correct": 0},
            "RI_END": {"total": 0, "correct": 0},
        }

        self.ws_server = VRWebSocketServer()

        # Admin controls
        self.subject_field = ft.TextField(label="N° de sujet", width=300)
        self.task_selector = ft.Dropdown(
            label="Type de tâche",
            width=300,
            options=[
                ft.dropdown.Option(key="recall", text="Tâche de rappel"),
                ft.dropdown.Option(key="rappel_immediat", text="Rappel immédiat"),
            ],
            value="recall",
        )
        self.start_btn = ft.Button(
            content=ft.Row([
                ft.Icon(ft.Icons.PLAY_ARROW),
                ft.Text("Démarrer la session")
            ], tight=True),
            on_click=self.on_start,
            style=ft.ButtonStyle(padding=20),
        )
        self.info_bar = ft.Text("")

        # Admin panel (visible before session)
        self.admin_panel = ft.Container(
            content=ft.Column(
                [
                    ft.Text("Configuration de la session", size=28, weight=ft.FontWeight.BOLD),
                    ft.Container(height=30),
                    self.subject_field,
                    ft.Container(height=10),
                    self.task_selector,
                    ft.Container(height=20),
                    self.start_btn,
                    ft.Container(height=20),
                    self.info_bar,
                ],
                horizontal_alignment=ft.CrossAxisAlignment.CENTER,
            ),
            alignment=ft.Alignment(0, 0),
            expand=True,
            padding=40,
        )

        # Fullscreen question display (visible during session) - ONLY the question content
        self.question_display = ft.Container(
            content=ft.Column([]),
            bgcolor=ft.Colors.GREY_200,
            alignment=ft.Alignment(0, 0),
            expand=True,
            visible=False,
        )

        # Splash screen - Logos
        self.splash_logos = ft.Container(
            content=ft.Column([
                ft.Row([
                    ft.Image(src=f"{ASSETS_DIR}/Logo_NP.jpg", height=150, fit="contain"),
                    ft.Container(width=80),
                    ft.Image(src=f"{ASSETS_DIR}/Logo_IRBA.jpeg", height=150, fit="contain"),
                ], alignment=ft.MainAxisAlignment.CENTER),
            ], alignment=ft.MainAxisAlignment.CENTER, horizontal_alignment=ft.CrossAxisAlignment.CENTER),
            bgcolor=ft.Colors.WHITE,
            alignment=ft.Alignment(0, 0),
            expand=True,
            opacity=1.0,
            animate_opacity=ft.Animation(800, ft.AnimationCurve.EASE_IN_OUT),
        )
        
        # Splash screen - HARMORYC title
        self.splash_title = ft.Container(
            content=ft.Column([
                ft.Text(
                    "HARMORYC",
                    size=72,
                    weight=ft.FontWeight.BOLD,
                    color=ft.Colors.BLUE_800,
                    text_align=ft.TextAlign.CENTER,
                ),
                ft.Text(
                    "Questionnaire Mémoire Épisodique",
                    size=24,
                    color=ft.Colors.GREY_600,
                    text_align=ft.TextAlign.CENTER,
                ),
            ], alignment=ft.MainAxisAlignment.CENTER, horizontal_alignment=ft.CrossAxisAlignment.CENTER, spacing=20),
            bgcolor=ft.Colors.WHITE,
            alignment=ft.Alignment(0, 0),
            expand=True,
            opacity=0.0,
            animate_opacity=ft.Animation(800, ft.AnimationCurve.EASE_IN_OUT),
            visible=False,
        )
        
        # Splash container (stacks logos and title)
        self.splash_screen = ft.Stack(
            [self.splash_logos, self.splash_title],
            expand=True,
        )

        # Navigation drawer for jumping to questions
        self.nav_drawer_content = ft.Column([], scroll=ft.ScrollMode.AUTO, spacing=2)
        self.nav_drawer = ft.Container(
            content=ft.Column([
                ft.Container(
                    content=ft.Text("Navigation", size=18, weight=ft.FontWeight.BOLD),
                    padding=15,
                    bgcolor=ft.Colors.BLUE_700,
                ),
                ft.Container(
                    content=self.nav_drawer_content,
                    padding=10,
                    expand=True,
                ),
            ], spacing=0),
            width=280,
            bgcolor=ft.Colors.WHITE,
            border=ft.border.only(right=ft.BorderSide(1, ft.Colors.GREY_400)),
            visible=False,
        )

        # High-resolution timer loop
        self._timer_task = self.page.run_task(self._timer_loop)

        # Keyboard shortcut handler
        self.page.on_keyboard_event = self._on_keyboard

        # Main app container (hidden during splash)
        self.main_app = ft.Row(
            [
                self.nav_drawer,
                ft.Stack(
                    [
                        self.admin_panel,
                        self.question_display,
                    ],
                    expand=True,
                ),
            ],
            expand=True,
            spacing=0,
            opacity=0.0,
            animate_opacity=ft.Animation(800, ft.AnimationCurve.EASE_IN_OUT),
            visible=False,
        )

        # Layout - stack splash and main app
        self.page.add(
            ft.Stack(
                [
                    self.splash_screen,
                    self.main_app,
                ],
                expand=True,
            )
        )
        self.page.update()
        
        # Start splash sequence and apply fullscreen
        self.page.run_task(self._splash_sequence)

    async def _splash_sequence(self):
        """Run the splash screen sequence with animations."""
        # Apply fullscreen first
        await asyncio.sleep(0.1)
        self.page.window_full_screen = True
        self.page.update()
        
        # Phase 1: Show logos for 3 seconds
        await asyncio.sleep(3.0)
        
        # Phase 2: Fade out logos, fade in HARMORYC
        self.splash_logos.opacity = 0.0
        self.splash_title.visible = True
        self.splash_title.opacity = 1.0
        self.page.update()
        
        # Wait for animation + display time
        await asyncio.sleep(2.5)
        
        # Phase 3: Fade out title, fade in main app
        self.splash_title.opacity = 0.0
        self.main_app.visible = True
        self.main_app.opacity = 1.0
        self.page.update()
        
        # Wait for animation to complete, then hide splash
        await asyncio.sleep(1.0)
        self.splash_screen.visible = False
        self.page.update()

    # ---------------------------
    # Event Handlers
    # ---------------------------
    def on_start(self, _):
        if self.running:
            return
        subject = (self.subject_field.value or "").strip()
        if not subject:
            self.toast_error("Veuillez saisir le N° de sujet.")
            return
        self.subject_id = subject
        self.task_mode = self.task_selector.value or "recall"

        # Load data
        try:
            ensure_sample_experiment_exists()
            with open(EXPERIMENT_FILE, "r", encoding="utf-8") as f:
                self.data = json.load(f)
        except Exception:
            self.data = {}

        # V2 assets override (primary data source)
        self.data = self._load_v2_assets()
        self.experiment_name = str(self.data.get("experiment_name", "MemoireEpisodique"))
        self.task_queue = self._build_tasks()
        if not self.task_queue:
            self.toast_error("Aucune tâche à exécuter.")
            return

        # Reset session state
        self.session_id = str(uuid.uuid4())
        self.session_started_at = now_iso()
        self.session_ended_at = None
        self.session_start_perf = time.perf_counter()
        self.current_task_index = -1
        self.records.clear()
        for k in self.stage_scores:
            self.stage_scores[k] = {"total": 0, "correct": 0}

        self.running = True
        self.task_active = False
        self.intermission_active = False
        self.default_image_counter = 0

        # Switch to session mode
        self.admin_panel.visible = False
        self.question_display.visible = True

        # Start WebSocket server (silent fail)
        try:
            self.ws_server.start()
        except Exception:
            pass

        self._go_to_next_task()
        self.page.update()

    def on_stop(self):
        self._finalize_current_task(timeout=True, force=False)
        if not self.running:
            return

        self.running = False
        self.task_active = False
        self.intermission_active = False

        self.session_ended_at = now_iso()

        try:
            vr_msgs = self.ws_server.drain_messages()
        except Exception:
            vr_msgs = []
        try:
            self.ws_server.stop()
        except Exception:
            pass

        self._save_session(vr_msgs)

        # Close navigation drawer and return to admin panel
        self.nav_drawer_open = False
        self.nav_drawer.visible = False
        self.question_display.visible = False
        self.admin_panel.visible = True
        self.subject_field.value = ""
        self.info_bar.value = "Session terminée et sauvegardée."
        self.page.update()

    # ---------------------------
    # Timer loop
    # ---------------------------
    async def _timer_loop(self):
        while True:
            await asyncio.sleep(0.05)
            self._timer_tick()

    def _timer_tick(self):
        if not self.running:
            return

        # Handle intermission
        if self.intermission_active and self.intermission_start_perf is not None:
            elapsed_ms = (time.perf_counter() - self.intermission_start_perf) * 1000.0
            if elapsed_ms >= INTER_DELAY_MS:
                self.intermission_active = False
                self.intermission_start_perf = None
                self._go_to_next_task()
            return

        # Handle task timeout
        if self.task_active and self.task_start_perf is not None:
            elapsed_ms = (time.perf_counter() - self.task_start_perf) * 1000.0
            if elapsed_ms >= self.task_duration_ms:
                self._finalize_current_task(timeout=True)

    # ---------------------------
    # Core flow
    # ---------------------------
    def _build_tasks(self) -> List[Dict[str, Any]]:
        rand = self._get_rng()
        tasks: List[Dict[str, Any]] = []

        rooms = list(self.data.get("rooms") or [])
        familiar = list(self.data.get("familiar_objects") or [])
        new_objects = list(self.data.get("new_objects") or [])
        rappel_immediat_trials = list(self.data.get("rappel_immediat_trials") or [])
        step5_trials = list(self.data.get("step5_trials") or [])
        if self.task_mode == "rappel_immediat":
            balanced_ri = self._pick_balanced_trials(rappel_immediat_trials, rand)
            rand.shuffle(balanced_ri)
            for t in balanced_ri:
                tasks.append({
                    "task_id": t.get("id"),
                    "kind": "rappel_immediat",
                    "stage": "RI",
                    "image": t.get("image"),
                    "is_correct": bool(t.get("is_correct")),
                    "duration_ms": RAPPEL_IMMEDIAT_DURATION_MS,
                })
            return tasks

        # Learning order (used for IIA/IV)
        rooms_learning = list(rooms)
        rand.shuffle(rooms_learning)
        self.learning_room_order = [r.get("id") for r in rooms_learning if r.get("id")]

        # IIA order (must differ from learning order)
        rooms_iia = list(rooms)
        for _ in range(5):
            rand.shuffle(rooms_iia)
            if [r.get("id") for r in rooms_iia] != self.learning_room_order:
                break
        if [r.get("id") for r in rooms_iia] == self.learning_room_order and rooms_iia:
            rooms_iia = rooms_iia[1:] + rooms_iia[:1]
        self.iia_room_order = [r.get("id") for r in rooms_iia if r.get("id")]

        # Étape I (reconnaissance d'objets)
        step1_objects = list(familiar) + list(new_objects)
        rand.shuffle(step1_objects)
        for idx, obj in enumerate(step1_objects, start=1):
            tasks.append({
                "task_id": f"step1_{obj.get('id')}_{idx}",
                "kind": "step1",
                "stage": "I",
                "object": obj,
                "duration_ms": STEP1_DURATION_MS,
            })

        # Étape III (jour/nuit par salle)
        rooms_for_step3 = list(rooms)
        rand.shuffle(rooms_for_step3)
        for room in rooms_for_step3:
            tasks.append({
                "task_id": f"step3_{room.get('id')}",
                "kind": "step3_daynight",
                "stage": "III",
                "room": room,
                "duration_ms": STEP3_DURATION_MS,
            })

        # Étape IV (ordre des salles)
        for room in rooms_for_step3:
            rid = room.get("id")
            correct_order = None
            if rid in self.learning_room_order:
                correct_order = self.learning_room_order.index(rid) + 1
            tasks.append({
                "task_id": f"step4_{rid}",
                "kind": "step4_order",
                "stage": "IV",
                "room": room,
                "correct_order": correct_order,
                "duration_ms": STEP4_DURATION_MS,
            })

        # Étape V (rappel tardif inversé)
        step5_trials = self._pick_balanced_trials(step5_trials, rand)
        rand.shuffle(step5_trials)
        for t in step5_trials:
            tasks.append({
                "task_id": f"step5_{t.get('id')}",
                "kind": "step5",
                "stage": "V",
                "image": t.get("image"),
                "is_correct": not bool(t.get("is_correct")),
                "duration_ms": STEP4_DURATION_MS,
            })

        # Rappel immédiat final (même logique que RI)
        final_ri = self._pick_balanced_trials(rappel_immediat_trials, rand)
        rand.shuffle(final_ri)
        for t in final_ri:
            tasks.append({
                "task_id": f"ri_end_{t.get('id')}",
                "kind": "rappel_immediat",
                "stage": "RI_END",
                "image": t.get("image"),
                "is_correct": bool(t.get("is_correct")),
                "duration_ms": RAPPEL_IMMEDIAT_DURATION_MS,
            })

        return tasks

    def _go_to_next_task(self):
        if not self.running:
            return
        self.current_task_index += 1
        if self.current_task_index >= len(self.task_queue):
            self.on_stop()
            return
        self.current_task = self.task_queue[self.current_task_index]
        self._show_task(self.current_task)

    def _get_question_text(self, task: Dict[str, Any]) -> str:
        """Get the question text based on task kind."""
        kind = task.get("kind")
        if kind == "step1":
            return "Avez-vous vu cet objet ?"
        elif kind == "step2_where":
            return "Dans quelle salle était cet objet ?"
        elif kind == "step2_spatial":
            return "Quelle position spatiale avez-vous vue ?"
        elif kind == "step3_daynight":
            return "Faisait-il jour ou nuit ?"
        elif kind == "step4_order":
            return "À quel moment cette salle a-t-elle été vue ? (1 = première salle vue, 10 = dernière salle vue)"
        elif kind == "step5":
            return "Cette scène est-elle correcte ?"
        elif kind == "rappel_immediat":
            return "Cette scène est-elle correcte ?"
        return ""

    def _show_task(self, task: Dict[str, Any]):
        self.task_active = False
        self.intermission_active = False
        self.task_started_at_iso = now_iso()
        self.task_start_perf = time.perf_counter()
        self.task_duration_ms = int(task.get("duration_ms", STEP1_DURATION_MS))
        self.task_active = True
        self.selected_choice_id = None
        self.choice_buttons.clear()

        kind = task.get("kind")
        content: List[ft.Control] = []

        # Always show the question text
        question_text = self._get_question_text(task)
        if question_text:
            content.append(ft.Text(question_text, size=28, weight=ft.FontWeight.W_600, text_align=ft.TextAlign.CENTER))
            content.append(ft.Container(height=20))

        if kind == "step1":
            obj = task["object"]
            content.append(self._image_or_default(obj.get("image"), image_type="object", height=IMAGE_HEIGHT_STEP1))
            content.append(ft.Container(height=30))
            content.append(self._button_row_yes_no())
            content.append(ft.Container(height=20))
            content.append(self._button_je_ne_sais_pas())

        elif kind == "step2_where":
            obj = task["object"]
            # Image de question un peu plus petite pour laisser plus de place aux réponses
            content.append(self._image_or_default(obj.get("image"), image_type="object", height=200))
            content.append(ft.Container(height=20))
            content.append(self._rooms_grid(task["rooms"]))
            content.append(ft.Container(height=15))
            content.append(self._button_je_ne_sais_pas())

        elif kind == "step2_spatial":
            obj = task.get("object", {})
            content.append(self._image_or_default(obj.get("image"), image_type="object", height=240))
            content.append(ft.Container(height=15))
            content.append(self._button_row_image_choices(task.get("choices") or []))
            content.append(ft.Container(height=15))
            content.append(self._button_je_ne_sais_pas())

        elif kind == "step3_daynight":
            room = task.get("room", {})
            content.append(self._image_or_default(room.get("image"), image_type="room", height=IMAGE_HEIGHT_STEP3))
            content.append(ft.Container(height=20))
            content.append(self._button_row_timing())
            content.append(ft.Container(height=15))
            content.append(self._button_je_ne_sais_pas())

        elif kind == "step4_order":
            room = task.get("room", {})
            content.append(self._image_or_default(room.get("image"), image_type="room", height=IMAGE_HEIGHT_STEP4))
            content.append(ft.Container(height=20))
            content.append(self._button_row_numbers(1, 10))
            content.append(ft.Container(height=15))
            content.append(self._button_je_ne_sais_pas())

        elif kind == "rappel_immediat":
            content.append(self._image_or_default(task.get("image"), image_type="object", height=IMAGE_HEIGHT_RAPPEL))
            content.append(ft.Container(height=30))
            content.append(self._button_row_correct_incorrect())
            content.append(ft.Container(height=15))
            content.append(self._button_je_ne_sais_pas())

        elif kind == "step5":
            content.append(self._image_or_default(task.get("image"), image_type="object", height=IMAGE_HEIGHT_STEP4))
            content.append(ft.Container(height=30))
            content.append(self._button_row_correct_incorrect())
            content.append(ft.Container(height=15))
            content.append(self._button_je_ne_sais_pas())

        else:
            content.append(ft.Text(f"Tâche inconnue: {kind}", color=ft.Colors.RED, size=24))

        # Update fullscreen question display - ONLY the content, nothing else
        self.question_display.content = ft.Column(
            content,
            spacing=0,
            horizontal_alignment=ft.CrossAxisAlignment.CENTER,
            alignment=ft.MainAxisAlignment.CENTER,
            expand=True,
        )
        self.question_display.update()
        
        # Update navigation drawer if open
        if self.nav_drawer_open:
            self._build_nav_drawer_content()
        
        self.page.update()

    # ---------------------------
    # Rendering helpers
    # ---------------------------
    def _get_default_object_image(self) -> str:
        """Get alternating default object image."""
        self.default_image_counter += 1
        if self.default_image_counter % 2 == 0:
            return f"{ASSETS_DIR}/Default_Object_2.jpg"
        return f"{ASSETS_DIR}/Default_Object_1.jpg"

    def _get_default_room_image(self) -> str:
        """Get alternating default room image."""
        self.default_image_counter += 1
        if self.default_image_counter % 2 == 0:
            return f"{ASSETS_DIR}/Default_Room_2.jpg"
        return f"{ASSETS_DIR}/Default_Room_1.jpg"

    def _as_asset_path(self, path: Path) -> str:
        return path.as_posix()

    def _list_images(self, directory: Path) -> List[str]:
        if not directory.exists():
            return []
        images = []
        for p in sorted(directory.rglob("*")):
            if p.is_file() and p.suffix.lower() in IMAGE_EXTS:
                images.append(self._as_asset_path(p))
        return images

    def _build_trials_from_dirs(self, correct_dir: Path, incorrect_dir: Path) -> List[Dict[str, Any]]:
        trials: List[Dict[str, Any]] = []
        for img in self._list_images(correct_dir):
            trials.append({
                "id": f"correct_{Path(img).stem}",
                "image": img,
                "is_correct": True,
            })
        for img in self._list_images(incorrect_dir):
            trials.append({
                "id": f"incorrect_{Path(img).stem}",
                "image": img,
                "is_correct": False,
            })
        return trials

    def _first_image(self, directory: Path) -> Optional[str]:
        images = self._list_images(directory)
        return images[0] if images else None

    def _extract_index(self, text: str) -> Optional[int]:
        match = re.search(r"(\\d+)", text)
        if not match:
            return None
        return int(match.group(1))

    def _get_rng(self) -> random.Random:
        if RANDOM_MODE == "per_session":
            seed = f"{self.experiment_name}-{self.subject_id}-{self.session_id}"
        else:
            seed = RANDOM_SEED
        return random.Random(seed)

    def _pick_balanced_trials(self, trials: List[Dict[str, Any]], rand: random.Random, target_each: int = 5) -> List[Dict[str, Any]]:
        correct = [t for t in trials if bool(t.get("is_correct"))]
        incorrect = [t for t in trials if not bool(t.get("is_correct"))]
        rand.shuffle(correct)
        rand.shuffle(incorrect)
        if len(correct) >= target_each and len(incorrect) >= target_each:
            return correct[:target_each] + incorrect[:target_each]
        return correct + incorrect

    def _load_v2_assets(self) -> Dict[str, Any]:
        base = Path(ASSETS_DIR) / V2_ASSETS_SUBDIR
        rooms_dir = base / "Start_Room"
        familiar_dir = base / "Objets" / "Objets_familiers (OF)"
        new_dir = base / "Objets" / "Nouveaux_objets (NO)"
        rappel_dir = base / "Rappel_immediat"
        rappel_correct_dir = rappel_dir / "Salles_correctes"
        rappel_incorrect_dir = rappel_dir / "Salles_incorrectes"
        iib_dir = base / "EtapesIIB"
        iib_familiar_dir = iib_dir / "OF"
        iib_new_dir = iib_dir / "NO"
        step5_dir = base / "EtapesV"
        step5_correct_dir = step5_dir / "Salles correctes"
        step5_incorrect_dir = step5_dir / "Salles incorrectes"

        rooms = []
        room_dirs = [d for d in rooms_dir.iterdir()] if rooms_dir.exists() else []
        room_dirs = [d for d in room_dirs if d.is_dir()]
        room_dirs.sort(key=lambda d: self._extract_index(d.name) or 0)
        if not room_dirs:
            room_dirs = [Path(f"Room{i}") for i in range(1, 11)]
        for idx, d in enumerate(room_dirs, start=1):
            room_index = self._extract_index(d.name) or idx
            room_id = f"room{room_index}"
            room_name = f"Salle {room_index}"
            room_image = self._first_image(d if d.exists() else rooms_dir / f"Room{room_index}")
            rooms.append({
                "id": room_id,
                "name": room_name,
                "image": room_image,
                "timing": "jour" if room_index <= 5 else "nuit",
            })

        familiar_objects = []
        for i in range(1, 11):
            obj_dir = familiar_dir / f"OF{i}"
            familiar_objects.append({
                "id": f"OF{i}",
                "name": f"Objet familier {i}",
                "image": self._first_image(obj_dir),
                "is_familiar": True,
                "room_id": f"room{i}",
                "timing": "jour" if i <= 5 else "nuit",
            })

        new_objects = []
        for i in range(1, 11):
            obj_dir = new_dir / f"NO{i}"
            new_objects.append({
                "id": f"NO{i}",
                "name": f"Nouvel objet {i}",
                "image": self._first_image(obj_dir),
                "is_familiar": False,
                "room_id": None,
                "timing": None,
            })

        rappel_immediat_trials = []
        for img in self._list_images(rappel_correct_dir):
            rappel_immediat_trials.append({
                "id": f"ri_correct_{Path(img).stem}",
                "image": img,
                "is_correct": True,
            })
        for img in self._list_images(rappel_incorrect_dir):
            rappel_immediat_trials.append({
                "id": f"ri_incorrect_{Path(img).stem}",
                "image": img,
                "is_correct": False,
            })

        def _position_images(
            obj_dir: Path,
            good_prefix: str,
            bad_prefixes: List[str],
            allow_good_fallback: bool = True,
        ) -> Tuple[Optional[str], Optional[str], Optional[str]]:
            if not obj_dir.exists():
                return None, None, None
            subdirs = [d for d in obj_dir.iterdir() if d.is_dir()]
            good_dirs = [d for d in subdirs if d.name.lower().startswith(good_prefix.lower())]
            bad_dirs = [d for d in subdirs if any(d.name.lower().startswith(p.lower()) for p in bad_prefixes)]
            bad_dirs.sort(key=lambda d: d.name.lower())
            good_img = self._first_image(good_dirs[0]) if good_dirs else None
            bad_img = self._first_image(bad_dirs[0]) if bad_dirs else None
            alt_bad_img = self._first_image(bad_dirs[1]) if len(bad_dirs) > 1 else None
            all_imgs = self._list_images(obj_dir)
            if allow_good_fallback and not good_img and all_imgs:
                good_img = all_imgs[0]
            if not bad_img and len(all_imgs) > 1:
                bad_img = all_imgs[1]
            if not bad_img:
                bad_img = good_img
            if not alt_bad_img and len(all_imgs) > 2:
                alt_bad_img = all_imgs[2]
            if not alt_bad_img:
                alt_bad_img = bad_img
            return good_img, bad_img, alt_bad_img

        iib_positions: Dict[str, Dict[str, Optional[str]]] = {}
        for i in range(1, 11):
            obj_id = f"OF{i}"
            good_img, bad_img, alt_bad_img = _position_images(
                iib_familiar_dir / obj_id,
                good_prefix="Position_OF_OK",
                bad_prefixes=["Position_OF_Wrong"],
            )
            iib_positions[obj_id] = {"good": good_img, "bad": bad_img, "alt_bad": alt_bad_img}
        for i in range(1, 11):
            obj_id = f"NO{i}"
            good_img, bad_img, alt_bad_img = _position_images(
                iib_new_dir / obj_id,
                good_prefix="Position_NO_OK",
                bad_prefixes=["Position_NO_Wrong1", "Position_NO_Wrong2"],
                allow_good_fallback=False,
            )
            iib_positions[obj_id] = {"good": good_img, "bad": bad_img, "alt_bad": alt_bad_img}

        return {
            "experiment_name": "HARMORYC_V2",
            "rooms": rooms,
            "familiar_objects": familiar_objects,
            "new_objects": new_objects,
            "rappel_immediat_trials": rappel_immediat_trials,
            "step5_trials": self._build_trials_from_dirs(step5_correct_dir, step5_incorrect_dir),
            "iib_positions": iib_positions,
        }

    def _rooms_by_ids(self, room_ids: List[str], rooms: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        lookup = {r.get("id"): r for r in rooms}
        return [lookup[rid] for rid in room_ids if rid in lookup]

    def _image_or_default(self, src: Optional[str], image_type: str = "object", height: int = 300, expand: bool = False) -> ft.Control:
        """Display image or use default based on type (object/room)."""
        if not src or not Path(src).exists():
            if image_type == "room":
                src = self._get_default_room_image()
            else:
                src = self._get_default_object_image()
        return ft.Image(
            src=str(src),
            height=None if expand else height,
            expand=expand,
            fit="contain",
            error_content=ft.Text("Image introuvable", size=20),
        )

    def _button_row_yes_no(self) -> ft.Control:
        btn_yes = ft.FilledButton(
            content=ft.Text("Oui", size=20),
            width=200,
            height=60,
            on_click=lambda _: self._on_choice("yes"),
        )
        btn_no = ft.OutlinedButton(
            content=ft.Text("Non", size=20),
            width=200,
            height=60,
            on_click=lambda _: self._on_choice("no"),
        )
        self.choice_buttons["yes"] = btn_yes
        self.choice_buttons["no"] = btn_no
        return ft.Row(
            [btn_yes, ft.Container(width=40), btn_no],
            alignment=ft.MainAxisAlignment.CENTER,
        )

    def _button_row_correct_incorrect(self) -> ft.Control:
        btn_correct = ft.FilledButton(
            content=ft.Text("Correct", size=20),
            width=200,
            height=60,
            on_click=lambda _: self._on_choice("correct"),
        )
        btn_incorrect = ft.OutlinedButton(
            content=ft.Text("Incorrect", size=20),
            width=200,
            height=60,
            on_click=lambda _: self._on_choice("incorrect"),
        )
        self.choice_buttons["correct"] = btn_correct
        self.choice_buttons["incorrect"] = btn_incorrect
        return ft.Row(
            [btn_correct, ft.Container(width=40), btn_incorrect],
            alignment=ft.MainAxisAlignment.CENTER,
        )

    def _button_row_timing(self) -> ft.Control:
        """Display day/night choices as clickable images."""
        card_jour = ft.Container(
            content=ft.Column([
                ft.Image(
                    src=f"{ASSETS_DIR}/Default_Day.jpg",
                    height=150,
                    fit="cover",
                    border_radius=ft.border_radius.only(top_left=12, top_right=12),
                    error_content=ft.Container(
                        content=ft.Icon(ft.Icons.WB_SUNNY, size=60, color=ft.Colors.ORANGE_400),
                        height=150,
                        alignment=ft.Alignment(0, 0),
                        bgcolor=ft.Colors.AMBER_100,
                    ),
                ),
                ft.Container(
                    content=ft.Text("Jour", size=18, weight=ft.FontWeight.W_500, text_align=ft.TextAlign.CENTER),
                    padding=10,
                ),
            ], spacing=0, horizontal_alignment=ft.CrossAxisAlignment.CENTER),
            width=180,
            bgcolor=ft.Colors.WHITE,
            border=ft.border.all(2, ft.Colors.GREY_400),
            border_radius=12,
            ink=True,
            on_click=lambda _: self._on_choice("jour"),
        )
        
        card_nuit = ft.Container(
            content=ft.Column([
                ft.Image(
                    src=f"{ASSETS_DIR}/Default_Night.jpg",
                    height=150,
                    fit="cover",
                    border_radius=ft.border_radius.only(top_left=12, top_right=12),
                    error_content=ft.Container(
                        content=ft.Icon(ft.Icons.NIGHTLIGHT_ROUND, size=60, color=ft.Colors.INDIGO_400),
                        height=150,
                        alignment=ft.Alignment(0, 0),
                        bgcolor=ft.Colors.INDIGO_100,
                    ),
                ),
                ft.Container(
                    content=ft.Text("Nuit", size=18, weight=ft.FontWeight.W_500, text_align=ft.TextAlign.CENTER),
                    padding=10,
                ),
            ], spacing=0, horizontal_alignment=ft.CrossAxisAlignment.CENTER),
            width=180,
            bgcolor=ft.Colors.WHITE,
            border=ft.border.all(2, ft.Colors.GREY_400),
            border_radius=12,
            ink=True,
            on_click=lambda _: self._on_choice("nuit"),
        )
        
        self.choice_buttons["jour"] = card_jour
        self.choice_buttons["nuit"] = card_nuit
        return ft.Row(
            [card_jour, ft.Container(width=40), card_nuit],
            alignment=ft.MainAxisAlignment.CENTER,
        )

    def _button_row_numbers(self, start: int, end: int) -> ft.Control:
        buttons = []
        for i in range(start, end + 1):
            btn = ft.OutlinedButton(
                content=ft.Text(str(i), size=16),
                width=70,
                height=50,
                on_click=lambda _, v=str(i): self._on_choice(v),
            )
            self.choice_buttons[str(i)] = btn
            buttons.append(btn)
        return ft.Row(
            buttons,
            alignment=ft.MainAxisAlignment.CENTER,
            wrap=True,
            spacing=10,
            run_spacing=10,
        )

    def _button_row_image_choices(self, choices: List[Dict[str, Any]]) -> ft.Control:
        cards = []
        for idx, choice in enumerate(choices):
            cid = choice.get("id", f"choice_{idx}")
            img = choice.get("image")
            card = ft.Container(
                content=ft.Column([
                    self._image_or_default(img, image_type="object", height=300),
                    ft.Container(
                        content=ft.Text(f"Image {idx + 1}", size=14, weight=ft.FontWeight.W_500),
                        padding=8,
                        alignment=ft.Alignment(0, 0),
                    ),
                ], spacing=0, horizontal_alignment=ft.CrossAxisAlignment.CENTER),
                width=220,
                bgcolor=ft.Colors.WHITE,
                border=ft.border.all(2, ft.Colors.GREY_400),
                border_radius=12,
                ink=True,
                on_click=lambda _, r=cid: self._on_choice(r),
            )
            self.choice_buttons[cid] = card
            cards.append(card)
        return ft.Row(
            cards,
            alignment=ft.MainAxisAlignment.CENTER,
            spacing=30,
        )

    def _button_je_ne_sais_pas(self) -> ft.Control:
        btn = ft.OutlinedButton(
            content=ft.Text("Je ne sais pas", size=18),
            width=240,
            height=50,
            on_click=lambda _: self._on_choice("je_ne_sais_pas"),
        )
        self.choice_buttons["je_ne_sais_pas"] = btn
        return btn

    def _button_row_rooms(self, rooms: List[Dict[str, Any]]) -> ft.Control:
        """Display 2 room choices as clickable images (for step3_where)."""
        cards = []
        for idx, room in enumerate(rooms):
            rid = room.get("id", "")
            rname = room.get("name", rid)
            rimage = room.get("image", "")
            
            # Use default room image if not available
            if not rimage or not Path(rimage).exists():
                rimage = f"{ASSETS_DIR}/Default_Room_{(idx % 2) + 1}.jpg"
            
            card = ft.Container(
                content=ft.Column([
                    ft.Image(
                        src=rimage,
                        height=140,
                        fit="cover",
                        border_radius=ft.border_radius.only(top_left=12, top_right=12),
                        error_content=ft.Container(
                            content=ft.Icon(ft.Icons.MEETING_ROOM, size=50, color=ft.Colors.GREY_600),
                            height=140,
                            alignment=ft.Alignment(0, 0),
                            bgcolor=ft.Colors.GREY_200,
                        ),
                    ),
                    ft.Container(
                        content=ft.Text(rname, size=16, weight=ft.FontWeight.W_500, text_align=ft.TextAlign.CENTER),
                        padding=10,
                    ),
                ], spacing=0, horizontal_alignment=ft.CrossAxisAlignment.CENTER),
                width=180,
                bgcolor=ft.Colors.WHITE,
                border=ft.border.all(2, ft.Colors.GREY_400),
                border_radius=12,
                ink=True,
                on_click=lambda _, r=rid: self._on_choice(r),
            )
            self.choice_buttons[rid] = card
            cards.append(card)
        
        return ft.Row(
            cards,
            alignment=ft.MainAxisAlignment.CENTER,
            spacing=30,
        )

    def _rooms_grid(self, rooms: List[Dict[str, Any]]) -> ft.Control:
        """Display all 10 rooms as clickable image cards (for step2_where)."""
        ordered_rooms: List[Dict[str, Any]] = []
        lookup = {r.get("id"): r for r in rooms}
        for rid in ROOM_GRID_ORDER:
            if rid in lookup:
                ordered_rooms.append(lookup[rid])
        for room in rooms:
            if room not in ordered_rooms:
                ordered_rooms.append(room)

        cards = []
        for idx, room in enumerate(ordered_rooms):
            rid = room.get("id", "")
            rname = room.get("name", rid)
            rimage = room.get("image", "")
            
            # Use default room image if not available
            if not rimage or not Path(rimage).exists():
                rimage = f"{ASSETS_DIR}/Default_Room_{(idx % 2) + 1}.jpg"
            
            card = ft.Container(
                content=ft.Column([
                    # Image de salle plus grande, sans texte ni numéro sous l'image
                    ft.Image(
                        src=rimage,
                        height=170,
                        width=230,
                        fit="cover",
                        border_radius=ft.border_radius.only(top_left=8, top_right=8),
                        error_content=ft.Container(
                            content=ft.Icon(ft.Icons.MEETING_ROOM, size=40, color=ft.Colors.GREY_600),
                            height=120,
                            width=160,
                            alignment=ft.Alignment(0, 0),
                            bgcolor=ft.Colors.GREY_200,
                        ),
                    ),
                ], spacing=0, horizontal_alignment=ft.CrossAxisAlignment.CENTER),
                width=240,
                bgcolor=ft.Colors.WHITE,
                border=ft.border.all(1, ft.Colors.GREY_400),
                border_radius=8,
                ink=True,
                on_click=lambda _, r=rid: self._on_choice(r),
            )
            self.choice_buttons[rid] = card
            cards.append(card)

        first_row = cards[:5]
        second_row = cards[5:10]
        return ft.Column(
            [
                ft.Row(first_row, alignment=ft.MainAxisAlignment.CENTER, spacing=10),
                ft.Row(second_row, alignment=ft.MainAxisAlignment.CENTER, spacing=10),
            ],
            spacing=10,
        )

    # ---------------------------
    # Response handling
    # ---------------------------
    def _on_choice(self, choice_id: str):
        if not self.task_active:
            return
        self.selected_choice_id = choice_id
        self._finalize_current_task(timeout=False)

    def _finalize_current_task(self, timeout: bool, force: bool = True):
        if not self.task_active and not force:
            return
        task = self.current_task
        if not task or self.task_start_perf is None:
            return

        end_perf = time.perf_counter()
        elapsed_ms = (end_perf - self.task_start_perf) * 1000.0
        response_duration_ms = int(max(0.0, min(elapsed_ms, float(self.task_duration_ms))))
        actual_elapsed_ms = int(max(0.0, elapsed_ms))

        response_value = self.selected_choice_id
        if timeout and response_value is None:
            response_value = "ne_repond_pas"

        record: Dict[str, Any] = {
            "task_id": task.get("task_id"),
            "stage": task.get("stage"),
            "kind": task.get("kind"),
            "started_at": self.task_started_at_iso,
            "ended_at": now_iso(),
            "duration_ms": self.task_duration_ms,
            "actual_elapsed_ms": actual_elapsed_ms,
            "response": response_value,
            "timeout": timeout,
        }

        correct = self._evaluate_correctness(task, response_value, timeout)
        if correct is not None:
            record["is_correct"] = correct
            stage = str(task.get("stage"))
            if stage in self.stage_scores:
                self.stage_scores[stage]["total"] += 1
                if correct:
                    self.stage_scores[stage]["correct"] += 1

        if "object" in task:
            record["object"] = task["object"]
        if "image" in task:
            record["image"] = task.get("image")
        if "is_correct" in task:
            record["expected_is_correct"] = task.get("is_correct")
        if "room" in task:
            record["room"] = task.get("room")
        if "correct_order" in task:
            record["correct_order"] = task.get("correct_order")
        if "choices" in task:
            record["choices"] = task.get("choices")
        if "correct_room" in task:
            record["correct_room"] = task.get("correct_room")
            record["distractor_room"] = task.get("distractor_room")
        if "correct_timing" in task:
            record["correct_timing"] = task.get("correct_timing")

        self._annotate_record_metrics(task, response_value, record)

        self.records.append(record)

        # Inject Étape IIA / IIB tasks if needed (immediate follow-up)
        if task.get("kind") == "step1" and response_value == "yes":
            obj = task["object"]
            rooms = self.data.get("rooms") or []
            ordered_rooms = self._rooms_by_ids(ROOM_GRID_ORDER, rooms)
            for room in rooms:
                if room not in ordered_rooms:
                    ordered_rooms.append(room)
            iib_positions = self.data.get("iib_positions") or {}
            pos = iib_positions.get(obj.get("id"), {})
            good_img = pos.get("good")
            bad_img = pos.get("bad")
            alt_bad_img = pos.get("alt_bad")
            if bool(obj.get("is_familiar")):
                choices = [
                    {"id": "good", "image": good_img, "is_correct": True},
                    {"id": "bad", "image": bad_img, "is_correct": False},
                ]
            else:
                choices = [
                    {"id": "bad1", "image": bad_img, "is_correct": False},
                    {"id": "bad2", "image": alt_bad_img, "is_correct": False},
                ]
            random.shuffle(choices)
            insert_at = self.current_task_index + 1
            self.task_queue[insert_at:insert_at] = [
                {
                    "task_id": f"step2_where_{obj.get('id')}",
                    "kind": "step2_where",
                    "stage": "IIA",
                    "object": obj,
                    "rooms": ordered_rooms,
                    "duration_ms": STEP2_DURATION_MS,
                },
                {
                    "task_id": f"step2_spatial_{obj.get('id')}",
                    "kind": "step2_spatial",
                    "stage": "IIB",
                    "object": obj,
                    "choices": choices,
                    "duration_ms": STEP2_DURATION_MS,
                },
            ]

        # Reset current task flags
        self.task_active = False
        self.task_start_perf = None
        self.task_started_at_iso = None
        self.selected_choice_id = None
        self.choice_buttons.clear()

        # Intermission screen
        self._start_intermission()

    def _start_intermission(self):
        self.intermission_active = True
        self.intermission_start_perf = time.perf_counter()
        self.question_display.content = ft.Container(
            bgcolor=ft.Colors.GREY_200,
            expand=True,
        )
        self.question_display.update()
        self.page.update()

    def _evaluate_correctness(self, task: Dict[str, Any], response: Optional[str], timeout: bool) -> Optional[bool]:
        kind = task.get("kind")
        if kind == "step2_spatial":
            obj = task.get("object", {})
            if not bool(obj.get("is_familiar")):
                return None
        if response is None or response in ("je_ne_sais_pas", "ne_repond_pas"):
            return False
        if kind == "step1":
            obj = task["object"]
            is_familiar = bool(obj.get("is_familiar"))
            if response == "yes":
                return is_familiar
            if response == "no":
                return not is_familiar
        elif kind == "step2_where":
            obj = task.get("object", {})
            if not bool(obj.get("is_familiar")):
                return False
            return response == obj.get("room_id")
        elif kind == "step2_spatial":
            obj = task.get("object", {})
            if not bool(obj.get("is_familiar")):
                return None
            for choice in task.get("choices") or []:
                if choice.get("id") == response:
                    return bool(choice.get("is_correct"))
            return False
        elif kind == "step3_daynight":
            return response == task.get("room", {}).get("timing")
        elif kind == "step4_order":
            correct_order = task.get("correct_order")
            return response == str(correct_order)
        elif kind in ("step5", "rappel_immediat"):
            expected = bool(task.get("is_correct"))
            if response == "correct":
                return expected
            if response == "incorrect":
                return not expected
        return None

    def _annotate_record_metrics(self, task: Dict[str, Any], response: Optional[str], record: Dict[str, Any]) -> None:
        kind = task.get("kind")
        if response is None:
            record["response_category"] = "no_response"
            return
        if response == "ne_repond_pas":
            record["response_category"] = "no_response"
            return
        if response == "je_ne_sais_pas":
            record["response_category"] = "dont_know"
            return

        if kind == "step1":
            obj = task.get("object", {})
            is_familiar = bool(obj.get("is_familiar"))
            record["expected_seen"] = is_familiar
            if response == "yes":
                record["response_category"] = "hit" if is_familiar else "false_alarm"
            elif response == "no":
                record["response_category"] = "miss" if is_familiar else "correct_rejection"
            else:
                record["response_category"] = "other"
            return

        if kind == "step2_where":
            obj = task.get("object", {})
            record["expected_room_id"] = obj.get("room_id")
            record["response_room_id"] = response
            if not bool(obj.get("is_familiar")):
                record["response_category"] = "no_correct_answer"
            elif response == obj.get("room_id"):
                record["response_category"] = "correct_room"
            else:
                record["response_category"] = "wrong_room"
            return

        if kind == "step2_spatial":
            obj = task.get("object", {})
            record["response_choice_id"] = response
            if not bool(obj.get("is_familiar")):
                record["response_category"] = "no_correct_answer"
                return
            expected_choice = None
            for choice in task.get("choices") or []:
                if choice.get("is_correct"):
                    expected_choice = choice.get("id")
                    break
            record["expected_choice_id"] = expected_choice
            if response == expected_choice:
                record["response_category"] = "correct_position"
            else:
                record["response_category"] = "wrong_position"
            return

        if kind == "step3_daynight":
            record["expected_timing"] = task.get("room", {}).get("timing")
            record["response_timing"] = response
            if response == record["expected_timing"]:
                record["response_category"] = "correct_time"
            else:
                record["response_category"] = "wrong_time"
            return

        if kind == "step4_order":
            expected = task.get("correct_order")
            record["expected_order"] = expected
            record["response_order"] = response
            try:
                response_int = int(response)
            except Exception:
                response_int = None
            if expected is not None and response_int is not None:
                record["error_distance"] = abs(response_int - int(expected))
                record["response_category"] = "correct_order" if response_int == expected else "wrong_order"
            else:
                record["response_category"] = "wrong_order"
            return

        if kind in ("step5", "rappel_immediat"):
            record["expected_is_correct"] = bool(task.get("is_correct"))
            record["response_category"] = "correct_scene" if record.get("is_correct") else "wrong_scene"
            return

    # ---------------------------
    # Persistence
    # ---------------------------
    def _compute_metrics(self) -> Dict[str, Any]:
        metrics: Dict[str, Any] = {
            "step1": {"hit": 0, "miss": 0, "false_alarm": 0, "correct_rejection": 0, "dont_know": 0},
            "step2_where": {"correct_room": 0, "wrong_room": 0, "dont_know": 0, "no_correct_answer": 0},
            "step2_spatial": {"correct_position": 0, "wrong_position": 0, "dont_know": 0, "no_correct_answer": 0},
            "step3_daynight": {"correct_time": 0, "wrong_time": 0, "dont_know": 0},
            "step4_order": {"correct_order": 0, "wrong_order": 0, "dont_know": 0, "mean_abs_error": None},
            "step5": {"correct_scene": 0, "wrong_scene": 0, "dont_know": 0},
            "rappel_immediat": {"correct_scene": 0, "wrong_scene": 0, "dont_know": 0},
        }
        order_errors: List[int] = []

        for rec in self.records:
            kind = rec.get("kind")
            category = rec.get("response_category")
            if kind == "step1":
                if category in metrics["step1"]:
                    metrics["step1"][category] += 1
            elif kind == "step2_where":
                if category in metrics["step2_where"]:
                    metrics["step2_where"][category] += 1
            elif kind == "step2_spatial":
                if category in metrics["step2_spatial"]:
                    metrics["step2_spatial"][category] += 1
            elif kind == "step3_daynight":
                if category in metrics["step3_daynight"]:
                    metrics["step3_daynight"][category] += 1
            elif kind == "step4_order":
                if category in metrics["step4_order"]:
                    metrics["step4_order"][category] += 1
                if isinstance(rec.get("error_distance"), int):
                    order_errors.append(rec["error_distance"])
            elif kind == "step5":
                if category in metrics["step5"]:
                    metrics["step5"][category] += 1
            elif kind == "rappel_immediat":
                if category in metrics["rappel_immediat"]:
                    metrics["rappel_immediat"][category] += 1

        if order_errors:
            metrics["step4_order"]["mean_abs_error"] = sum(order_errors) / len(order_errors)
        return metrics

    def _load_all_sessions(self) -> List[Dict[str, Any]]:
        """Load all session JSON files from sessions directory."""
        if not HAS_EXCEL_SUPPORT:
            return []
        
        sessions_data = []
        sessions_path = Path(SESSIONS_DIR)
        if not sessions_path.exists():
            return sessions_data
        
        for json_file in sessions_path.glob("*.json"):
            if json_file.name == "example_session.json":
                continue  # Skip example file
            try:
                with open(json_file, "r", encoding="utf-8") as f:
                    data = json.load(f)
                    if "session" in data:
                        sessions_data.append(data)
            except Exception:
                # Skip corrupted files
                continue
        
        return sessions_data

    def _generate_summary_excel(self, sessions_data: List[Dict[str, Any]]) -> None:
        """Generate Excel summary file with one row per session."""
        if not HAS_EXCEL_SUPPORT or not sessions_data:
            return
        
        try:
            rows = []
            for session_data in sessions_data:
                session = session_data.get("session", {})
                scores = session.get("scores", {})
                metrics = session.get("metrics", {})
                
                # Calculate percentages
                def calc_percentage(stage_key: str) -> Optional[float]:
                    stage_scores = scores.get(stage_key, {})
                    total = stage_scores.get("total", 0)
                    correct = stage_scores.get("correct", 0)
                    if total > 0:
                        return (correct / total) * 100.0
                    return None
                
                # Calculate duration
                duration_minutes = None
                started_at = session.get("started_at")
                ended_at = session.get("ended_at")
                if started_at and ended_at:
                    try:
                        start_dt = datetime.fromisoformat(started_at.replace("Z", "+00:00"))
                        end_dt = datetime.fromisoformat(ended_at.replace("Z", "+00:00"))
                        duration_minutes = (end_dt - start_dt).total_seconds() / 60.0
                    except Exception:
                        pass
                
                # Extract metrics
                step1_metrics = metrics.get("step1", {})
                step2_where_metrics = metrics.get("step2_where", {})
                step2_spatial_metrics = metrics.get("step2_spatial", {})
                step3_metrics = metrics.get("step3_daynight", {})
                step4_metrics = metrics.get("step4_order", {})
                step5_metrics = metrics.get("step5", {})
                ri_metrics = metrics.get("rappel_immediat", {})
                
                row = {
                    "session_id": session.get("session_id", ""),
                    "subject_id": session.get("subject_id", ""),
                    "started_at": started_at,
                    "ended_at": ended_at,
                    "task_mode": session.get("task_mode", ""),
                    "experiment_name": session.get("experiment_name", ""),
                    # Scores bruts
                    "I_total": scores.get("I", {}).get("total", 0),
                    "I_correct": scores.get("I", {}).get("correct", 0),
                    "IIA_total": scores.get("IIA", {}).get("total", 0),
                    "IIA_correct": scores.get("IIA", {}).get("correct", 0),
                    "IIB_total": scores.get("IIB", {}).get("total", 0),
                    "IIB_correct": scores.get("IIB", {}).get("correct", 0),
                    "III_total": scores.get("III", {}).get("total", 0),
                    "III_correct": scores.get("III", {}).get("correct", 0),
                    "IV_total": scores.get("IV", {}).get("total", 0),
                    "IV_correct": scores.get("IV", {}).get("correct", 0),
                    "V_total": scores.get("V", {}).get("total", 0),
                    "V_correct": scores.get("V", {}).get("correct", 0),
                    "RI_total": scores.get("RI", {}).get("total", 0),
                    "RI_correct": scores.get("RI", {}).get("correct", 0),
                    # Pourcentages
                    "I_%": calc_percentage("I"),
                    "IIA_%": calc_percentage("IIA"),
                    "IIB_%": calc_percentage("IIB"),
                    "III_%": calc_percentage("III"),
                    "IV_%": calc_percentage("IV"),
                    "V_%": calc_percentage("V"),
                    "RI_%": calc_percentage("RI"),
                    # Métriques Étape I
                    "step1_hit": step1_metrics.get("hit", 0),
                    "step1_false_alarm": step1_metrics.get("false_alarm", 0),
                    "step1_miss": step1_metrics.get("miss", 0),
                    "step1_correct_rejection": step1_metrics.get("correct_rejection", 0),
                    # Métriques Étape IIA
                    "step2_where_correct_room": step2_where_metrics.get("correct_room", 0),
                    "step2_where_wrong_room": step2_where_metrics.get("wrong_room", 0),
                    # Métriques Étape IIB
                    "step2_spatial_correct_position": step2_spatial_metrics.get("correct_position", 0),
                    "step2_spatial_wrong_position": step2_spatial_metrics.get("wrong_position", 0),
                    # Métriques Étape III
                    "step3_correct_time": step3_metrics.get("correct_time", 0),
                    "step3_wrong_time": step3_metrics.get("wrong_time", 0),
                    # Métriques Étape IV
                    "step4_correct_order": step4_metrics.get("correct_order", 0),
                    "step4_mean_abs_error": step4_metrics.get("mean_abs_error"),
                    # Métriques Étape V
                    "step5_correct_scene": step5_metrics.get("correct_scene", 0),
                    "step5_wrong_scene": step5_metrics.get("wrong_scene", 0),
                    # Métriques RI
                    "rappel_immediat_correct_scene": ri_metrics.get("correct_scene", 0),
                    "rappel_immediat_wrong_scene": ri_metrics.get("wrong_scene", 0),
                    # Durée
                    "duration_minutes": duration_minutes,
                }
                rows.append(row)
            
            df = pd.DataFrame(rows)
            
            # Sort by started_at if available
            if "started_at" in df.columns:
                df["started_at_parsed"] = pd.to_datetime(df["started_at"], errors="coerce")
                df = df.sort_values("started_at_parsed", na_position="last")
                df = df.drop(columns=["started_at_parsed"])
            
            # Save to Excel
            excel_path = Path(SESSIONS_DIR) / "summary.xlsx"
            with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
                df.to_excel(writer, sheet_name="Sessions", index=False)
                
                # Format the worksheet
                worksheet = writer.sheets["Sessions"]
                
                # Header formatting
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")
                
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except Exception:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                
                # Format percentage columns
                for col_idx, col_name in enumerate(df.columns, start=1):
                    if col_name.endswith("_%"):
                        col_letter = get_column_letter(col_idx)
                        for row_idx in range(2, len(df) + 2):
                            cell = worksheet[f"{col_letter}{row_idx}"]
                            if cell.value is not None:
                                cell.number_format = "0.00"
                
        except Exception:
            # Silently fail - don't block session save
            pass

    def _generate_summary_graphics(self, sessions_data: List[Dict[str, Any]]) -> None:
        """Generate summary graphics from all sessions."""
        if not HAS_EXCEL_SUPPORT or len(sessions_data) < 1:
            return
        
        try:
            # Prepare data
            stages = ["I", "IIA", "IIB", "III", "IV", "V", "RI"]
            percentages_by_stage = {stage: [] for stage in stages}
            dates = []
            
            for session_data in sessions_data:
                session = session_data.get("session", {})
                scores = session.get("scores", {})
                
                # Calculate percentages
                for stage in stages:
                    stage_scores = scores.get(stage, {})
                    total = stage_scores.get("total", 0)
                    correct = stage_scores.get("correct", 0)
                    if total > 0:
                        percentages_by_stage[stage].append((correct / total) * 100.0)
                    else:
                        percentages_by_stage[stage].append(None)
                
                # Extract date
                started_at = session.get("started_at")
                if started_at:
                    try:
                        date_dt = datetime.fromisoformat(started_at.replace("Z", "+00:00"))
                        dates.append(date_dt)
                    except Exception:
                        dates.append(None)
                else:
                    dates.append(None)
            
            # Filter out None dates for time series
            valid_indices = [i for i, d in enumerate(dates) if d is not None]
            if len(valid_indices) < 1:
                dates_valid = None
            else:
                dates_valid = [dates[i] for i in valid_indices]
            
            # Create figure with subplots
            fig = plt.figure(figsize=(16, 12))
            plt.rcParams['font.size'] = 9
            
            # 1. Bar chart: Average scores by stage
            ax1 = plt.subplot(2, 2, 1)
            stage_means = []
            stage_stds = []
            stage_labels = []
            
            for stage in stages:
                pcts = [p for p in percentages_by_stage[stage] if p is not None]
                if pcts:
                    stage_means.append(sum(pcts) / len(pcts))
                    if len(pcts) > 1:
                        variance = sum((x - stage_means[-1]) ** 2 for x in pcts) / (len(pcts) - 1)
                        stage_stds.append(variance ** 0.5)
                    else:
                        stage_stds.append(0)
                    stage_labels.append(stage)
                else:
                    stage_means.append(0)
                    stage_stds.append(0)
                    stage_labels.append(stage)
            
            x_pos = range(len(stage_labels))
            bars = ax1.bar(x_pos, stage_means, yerr=stage_stds, capsize=5, color='steelblue', alpha=0.7)
            ax1.set_xlabel("Étape")
            ax1.set_ylabel("Pourcentage de réussite (%)")
            ax1.set_title("Scores moyens par étape")
            ax1.set_xticks(x_pos)
            ax1.set_xticklabels(stage_labels)
            ax1.set_ylim(0, 100)
            ax1.grid(True, alpha=0.3, axis='y')
            
            # 2. Line chart: Evolution over time
            ax2 = plt.subplot(2, 2, 2)
            if dates_valid and len(dates_valid) > 1:
                for stage in stages:
                    pcts_valid = [percentages_by_stage[stage][i] for i in valid_indices if percentages_by_stage[stage][i] is not None]
                    dates_stage = [dates_valid[i] for i in range(len(valid_indices)) if percentages_by_stage[stage][valid_indices[i]] is not None]
                    if len(dates_stage) > 0:
                        ax2.plot(dates_stage, pcts_valid, marker='o', label=stage, linewidth=2, markersize=4)
                
                ax2.set_xlabel("Date")
                ax2.set_ylabel("Pourcentage de réussite (%)")
                ax2.set_title("Évolution des scores dans le temps")
                ax2.legend(loc='best', fontsize=8)
                ax2.grid(True, alpha=0.3)
                ax2.set_ylim(0, 100)
                plt.setp(ax2.xaxis.get_majorticklabels(), rotation=45, ha='right')
            else:
                ax2.text(0.5, 0.5, "Données insuffisantes\npour l'évolution temporelle", 
                        ha='center', va='center', transform=ax2.transAxes)
                ax2.set_title("Évolution des scores dans le temps")
            
            # 3. Histograms: Distribution of performance
            ax3 = plt.subplot(2, 2, 3)
            # Combine all percentages for overall distribution
            all_pcts = []
            for stage in stages:
                all_pcts.extend([p for p in percentages_by_stage[stage] if p is not None])
            
            if all_pcts:
                ax3.hist(all_pcts, bins=20, range=(0, 100), color='steelblue', alpha=0.7, edgecolor='black')
                ax3.set_xlabel("Pourcentage de réussite (%)")
                ax3.set_ylabel("Nombre de sessions")
                ax3.set_title("Distribution globale des performances")
                ax3.grid(True, alpha=0.3, axis='y')
            else:
                ax3.text(0.5, 0.5, "Aucune donnée disponible", 
                        ha='center', va='center', transform=ax3.transAxes)
                ax3.set_title("Distribution globale des performances")
            
            # 4. Combined chart: Step I metrics
            ax4 = plt.subplot(2, 2, 4)
            step1_hits = []
            step1_false_alarms = []
            step1_misses = []
            step1_correct_rejections = []
            
            for session_data in sessions_data:
                metrics = session_data.get("session", {}).get("metrics", {})
                step1_metrics = metrics.get("step1", {})
                step1_hits.append(step1_metrics.get("hit", 0))
                step1_false_alarms.append(step1_metrics.get("false_alarm", 0))
                step1_misses.append(step1_metrics.get("miss", 0))
                step1_correct_rejections.append(step1_metrics.get("correct_rejection", 0))
            
            if step1_hits:
                x_pos_metrics = range(len(step1_hits))
                width = 0.2
                ax4.bar([x - 1.5*width for x in x_pos_metrics], step1_hits, width, label='Hits', color='green', alpha=0.7)
                ax4.bar([x - 0.5*width for x in x_pos_metrics], step1_false_alarms, width, label='Fausses alarmes', color='red', alpha=0.7)
                ax4.bar([x + 0.5*width for x in x_pos_metrics], step1_misses, width, label='Misses', color='orange', alpha=0.7)
                ax4.bar([x + 1.5*width for x in x_pos_metrics], step1_correct_rejections, width, label='Rejets corrects', color='blue', alpha=0.7)
                
                ax4.set_xlabel("Session")
                ax4.set_ylabel("Nombre")
                ax4.set_title("Métriques principales - Étape I")
                ax4.legend(loc='best', fontsize=8)
                ax4.set_xticks(x_pos_metrics)
                ax4.set_xticklabels([f"S{i+1}" for i in range(len(step1_hits))], rotation=45, ha='right')
                ax4.grid(True, alpha=0.3, axis='y')
            else:
                ax4.text(0.5, 0.5, "Aucune donnée disponible", 
                        ha='center', va='center', transform=ax4.transAxes)
                ax4.set_title("Métriques principales - Étape I")
            
            plt.tight_layout()
            
            # Save figure
            graphics_path = Path(SESSIONS_DIR) / "summary_graphics.png"
            plt.savefig(graphics_path, dpi=150, bbox_inches='tight')
            plt.close(fig)
            
        except Exception:
            # Silently fail - don't block session save
            pass

    def _save_session(self, vr_msgs: List[Dict[str, Any]]) -> None:
        start_perf = self.session_start_perf or time.perf_counter()
        vr_serialized: List[Dict[str, Any]] = []
        for m in vr_msgs:
            rp = float(m.get("received_perf", start_perf))
            relative_ms = int(max(0.0, (rp - start_perf) * 1000.0))
            vr_serialized.append({
                "received_at": m.get("received_at"),
                "relative_ms": relative_ms,
                "payload": m.get("payload"),
            })

        meta = {
            "session_id": self.session_id,
            "subject_id": self.subject_id,
            "task_mode": self.task_mode,
            "experiment_name": self.experiment_name,
            "started_at": self.session_started_at,
            "ended_at": self.session_ended_at,
            "platform": {
                "system": platform.system(),
                "release": platform.release(),
                "version": platform.version(),
                "machine": platform.machine(),
                "processor": platform.processor(),
                "python_version": platform.python_version(),
                "node": platform.node(),
            },
            "scores": self.stage_scores,
            "metrics": self._compute_metrics(),
        }

        out = {
            "session": meta,
            "tasks": self.records,
            "vr_messages": vr_serialized,
        }

        Path(SESSIONS_DIR).mkdir(parents=True, exist_ok=True)
        ts_for_fn = safe_filename((self.session_started_at or now_iso()).replace(":", "-"))
        pid_for_fn = safe_filename(self.subject_id)
        fn = f"{ts_for_fn}_{pid_for_fn}_{self.session_id}.json"
        out_path = Path(SESSIONS_DIR) / fn
        try:
            with open(out_path, "w", encoding="utf-8") as f:
                json.dump(out, f, indent=2, ensure_ascii=False)
        except Exception:
            pass
        
        # Generate Excel summary and graphics
        try:
            all_sessions = self._load_all_sessions()
            self._generate_summary_excel(all_sessions)
            self._generate_summary_graphics(all_sessions)
        except Exception:
            # Silently fail - don't block session save
            pass

    # ---------------------------
    # UX helpers
    # ---------------------------
    def toast_error(self, message: str) -> None:
        self.page.snack_bar = ft.SnackBar(
            ft.Text(message),
            bgcolor=ft.Colors.RED_400,
            action=ft.SnackBarAction(label="OK"),
        )
        self.page.snack_bar.open = True
        self.page.update()

    # ---------------------------
    # Navigation drawer
    # ---------------------------
    def _on_keyboard(self, e: ft.KeyboardEvent):
        """Handle keyboard shortcuts. Press Ctrl+N to toggle navigation drawer."""
        if e.key == "N" and self.running and (getattr(e, "ctrl", False) or getattr(e, "meta", False)):
            self._toggle_nav_drawer()

    def _toggle_nav_drawer(self):
        """Toggle the navigation drawer visibility."""
        self.nav_drawer_open = not self.nav_drawer_open
        self.nav_drawer.visible = self.nav_drawer_open
        if self.nav_drawer_open:
            self._build_nav_drawer_content()
        self.page.update()

    def _build_nav_drawer_content(self):
        """Build the navigation drawer content with all questions."""
        items: List[ft.Control] = []
        
        # Group tasks by stage
        stages = {"I": [], "IIA": [], "IIB": [], "III": [], "IV": [], "V": [], "RI": [], "RI_END": []}
        for idx, task in enumerate(self.task_queue):
            stage = task.get("stage", "?")
            if stage in stages:
                stages[stage].append((idx, task))

        stage_names = {
            "I": "Étape I - Objets",
            "IIA": "Étape IIA - Salles",
            "IIB": "Étape IIB - Position spatiale",
            "III": "Étape III - Jour/Nuit",
            "IV": "Étape IV - Ordre",
            "V": "Étape V - Rappel tardif",
            "RI": "Rappel immédiat",
            "RI_END": "Rappel immédiat (fin)",
        }

        for stage_key, stage_tasks in stages.items():
            if not stage_tasks:
                continue
            
            # Stage header
            items.append(ft.Container(
                content=ft.Text(stage_names.get(stage_key, stage_key), size=14, weight=ft.FontWeight.BOLD),
                padding=ft.padding.only(top=10, bottom=5, left=5),
            ))
            
            # Task items
            for idx, task in stage_tasks:
                is_current = idx == self.current_task_index
                is_done = idx < self.current_task_index
                
                # Determine label
                kind = task.get("kind", "")
                obj = task.get("object", {})
                label = f"Q{idx + 1}"
                if obj:
                    label = f"Q{idx + 1}: {obj.get('name', obj.get('id', ''))[:15]}"
                elif task.get("room"):
                    room = task.get("room") or {}
                    label = f"Q{idx + 1}: {room.get('name', room.get('id', ''))[:15]}"
                elif kind in ("step5", "rappel_immediat"):
                    label = f"Q{idx + 1}: Scène"
                
                # Style based on state
                bgcolor = ft.Colors.BLUE_100 if is_current else (ft.Colors.GREY_200 if is_done else None)
                text_color = ft.Colors.BLUE_900 if is_current else (ft.Colors.GREY_600 if is_done else ft.Colors.BLACK)
                icon = ft.Icons.CHECK_CIRCLE if is_done else (ft.Icons.PLAY_CIRCLE if is_current else ft.Icons.CIRCLE_OUTLINED)
                
                items.append(ft.Container(
                    content=ft.Row([
                        ft.Icon(icon, size=16, color=text_color),
                        ft.Text(label, size=12, color=text_color, expand=True),
                    ], spacing=8),
                    padding=ft.padding.symmetric(horizontal=10, vertical=6),
                    bgcolor=bgcolor,
                    border_radius=4,
                    on_click=lambda _, i=idx: self._jump_to_question(i),
                    ink=True,
                ))

        self.nav_drawer_content.controls = items
        self.nav_drawer_content.update()

    def _jump_to_question(self, index: int):
        """Jump directly to a specific question."""
        if not self.running or index < 0 or index >= len(self.task_queue):
            return
        
        # Finalize current task as skipped
        if self.task_active:
            self._finalize_current_task(timeout=True, force=True)
        
        # Jump to the target question
        self.current_task_index = index - 1  # Will be incremented by _go_to_next_task
        self.intermission_active = False
        self.intermission_start_perf = None
        self._go_to_next_task()
        
        # Update drawer to show new position
        self._build_nav_drawer_content()
        self.page.update()


def main(page: ft.Page):
    ExperimentApp(page)


if __name__ == "__main__":
    ft.run(main)
